import openpyxl
from openpyxl.styles import PatternFill

def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheet_name]
    return  sheet.max_row

def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return  sheet.max_column

def read_data(file,sheet_name,row,column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    cell = sheet.cell(row,column)
    return cell.value

def write_data(file,sheet_name,row,column,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    cell = sheet.cell(row,column)
    cell.value = data
    workbook.save(file)

def fill_green_color(file,sheet_name,row,column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    green_fill = PatternFill(fill_type="solid",start_color="60b212",end_color="60b212")
    cell = sheet.cell(row, column)
    cell.fill = green_fill
    workbook.save(file)

def fill_red_color(file,sheet_name,row,column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    red_fill = PatternFill(fill_type="solid",start_color="ff0000",end_color="ff0000")
    cell = sheet.cell(row, column)
    cell.fill = red_fill
    workbook.save(file)


